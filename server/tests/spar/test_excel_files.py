"""Tests for Excel file handling fixes.

This module tests:
1. MIME type detection for Excel files
2. Table naming for single-sheet vs multi-sheet Excel files
3. CSV compatibility
"""

from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest

# MIME Type Detection Tests


def test_xlsx_file_mime_type():
    """Test that .xlsx files are correctly identified."""
    from agent_platform.server.file_manager.utils import guess_mimetype

    # Create a minimal valid XLSX file (ZIP format with PK header)
    xlsx_bytes = b"PK\x03\x04"  # ZIP file header
    mime_type = guess_mimetype("test.xlsx", xlsx_bytes)
    assert mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_xls_file_mime_type():
    """Test that .xls files are correctly identified."""
    from agent_platform.server.file_manager.utils import guess_mimetype

    # Create a minimal XLS file header
    xls_bytes = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    mime_type = guess_mimetype("test.xls", xls_bytes)
    assert mime_type == "application/vnd.ms-excel"


def test_csv_file_mime_type():
    """Test that .csv files are correctly identified."""
    from agent_platform.server.file_manager.utils import guess_mimetype

    csv_bytes = b"name,age\nJohn,25"
    mime_type = guess_mimetype("test.csv", csv_bytes)
    # On windows the mime type may be "application/vnd.ms-excel" even for a .csv file.
    assert mime_type in ("application/vnd.ms-excel", "text/csv")


def test_xlsx_without_header():
    """Test that .xlsx extension takes precedence over content."""
    from agent_platform.server.file_manager.utils import guess_mimetype

    # Even with non-ZIP content, extension should be checked first
    random_bytes = b"random content"
    mime_type = guess_mimetype("data.xlsx", random_bytes)
    assert mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# Table Naming Tests


def test_csv_table_uses_filename():
    """Test that CSV files use the filename as table name."""
    from agent_platform.server.api.private_v2.data_connections import _create_table_info
    from agent_platform.server.data_frames.data_reader import CsvDataReader

    csv_bytes = b"name,age\nJohn,25\nJane,30"
    reader = CsvDataReader(csv_bytes)
    sheet = next(reader.iter_sheets())

    has_multiple_sheets = reader.has_multiple_sheets()
    table_info = _create_table_info(sheet, "test_data.csv", has_multiple_sheets)

    assert table_info.name == "test_data.csv"
    assert sheet.name is None
    assert has_multiple_sheets is False


def test_single_sheet_excel_uses_filename():
    """Test that single-sheet Excel files use filename as table name."""
    from agent_platform.server.api.private_v2.data_connections import _create_table_info
    from agent_platform.server.data_frames.data_reader import ExcelDataReader

    # Use the sample.xlsx file for this test (it has multiple sheets, but we'll test the logic)
    test_data_path = Path(__file__).parent.parent / "data_frames" / "test_file_data_readers"
    test_file = test_data_path / "sample.xlsx"

    if not test_file.exists():
        pytest.skip("Test Excel file not found")

    file_bytes = test_file.read_bytes()
    reader = ExcelDataReader(file_bytes)
    has_multiple_sheets = reader.has_multiple_sheets()

    # Test the logic: for a single sheet, it should use filename
    # We'll test with mock data since sample.xlsx has multiple sheets
    if has_multiple_sheets:
        # Use the first sheet but test single-sheet behavior
        sheet = next(reader.iter_sheets())
        table_info = _create_table_info(sheet, "report.xlsx", has_multiple_sheets=False)

        # With has_multiple_sheets=False, should use filename
        assert table_info.name == "report.xlsx"


def test_multi_sheet_excel_uses_sheet_names():
    """Test that multi-sheet Excel files use sheet names as table names."""
    from agent_platform.server.api.private_v2.data_connections import _create_table_info
    from agent_platform.server.data_frames.data_reader import ExcelDataReader

    # Use the correct test data directory
    test_data_path = Path(__file__).parent.parent / "data_frames" / "test_file_data_readers"
    test_file = test_data_path / "sample.xlsx"

    if not test_file.exists():
        pytest.skip("Test Excel file not found")

    file_bytes = test_file.read_bytes()
    reader = ExcelDataReader(file_bytes)
    has_multiple_sheets = reader.has_multiple_sheets()

    # For multi-sheet Excel, should use sheet names
    if has_multiple_sheets:
        sheets = list(reader.iter_sheets())
        assert len(sheets) > 1

        for sheet in sheets:
            table_info = _create_table_info(sheet, "workbook.xlsx", has_multiple_sheets)
            # Should use sheet name, not filename
            assert table_info.name == sheet.name
            assert table_info.name != "workbook.xlsx"


def test_table_naming_with_mock_sheets():
    """Test table naming logic with mocked sheet objects."""
    from agent_platform.server.api.private_v2.data_connections import _create_table_info

    # Mock a single-sheet scenario
    mock_sheet = MagicMock()
    mock_sheet.name = "Sheet1"
    mock_sheet.column_headers = ["name", "age"]
    mock_sheet.list_sample_rows.return_value = [["John", 25], ["Jane", 30]]

    # Single sheet - should use filename
    table_info = _create_table_info(mock_sheet, "data.xlsx", has_multiple_sheets=False)
    assert table_info.name == "data.xlsx"

    # Multiple sheets - should use sheet name
    table_info = _create_table_info(mock_sheet, "data.xlsx", has_multiple_sheets=True)
    assert table_info.name == "Sheet1"

    # CSV (sheet.name is None) - should use filename
    mock_sheet.name = None
    table_info = _create_table_info(mock_sheet, "data.csv", has_multiple_sheets=False)
    assert table_info.name == "data.csv"


# Excel Data Reader Logging Tests


def test_excel_reader_logs_warning_on_fallback(capsys):
    """Test that falling back to CSV logs a warning."""
    from agent_platform.server.data_frames.data_reader import ExcelDataReader

    # Invalid Excel data that will fallback to CSV
    csv_bytes = b"name,age\nJohn,25"
    reader = ExcelDataReader(csv_bytes)

    # Trigger the fallback by iterating sheets
    sheets = list(reader.iter_sheets())

    # Should successfully fall back and read as CSV
    assert len(sheets) == 1
    assert sheets[0].num_rows == 1

    # Check if warning was logged to stdout
    captured = capsys.readouterr()
    output = (captured.out + captured.err).lower()
    assert "failed to read" in output or "falling back" in output


# End-to-End Tests


@pytest.mark.asyncio
async def test_inspect_single_sheet_excel():
    """Test inspecting a single-sheet Excel file returns correct table name."""
    import openpyxl
    from fastapi import Request

    from agent_platform.server.api.private_v2.data_connections import (
        inspect_file_as_data_connection,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Sheet1"
        ws.append(["name", "age"])
        ws.append(["John", 25])
        ws.append(["Jane", 30])

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    # Mock request
    mock_request = MagicMock(spec=Request)
    mock_request.headers.get.return_value = "test_file.xlsx"
    mock_request.body = AsyncMock(return_value=excel_bytes)

    # Mock user
    mock_user = MagicMock()
    mock_user.user_id = "test-user"

    response = await inspect_file_as_data_connection(mock_request, mock_user)

    # Should have one table
    assert len(response.tables) == 1

    # Table name should be the filename, not "Sheet1"
    assert response.tables[0].name == "test_file.xlsx"

    # Should have correct columns
    column_names = [col.name for col in response.tables[0].columns]
    assert "name" in column_names
    assert "age" in column_names


# Helper Functions


def create_test_excel_file(sheet_data: dict[str, list[list]], filename: str = "test.xlsx") -> bytes:
    """Create a test Excel file with given data.

    Args:
        sheet_data: Dict of {sheet_name: [[row1], [row2], ...]}
        filename: Name for the file

    Returns:
        bytes: The Excel file content
    """
    import openpyxl

    wb = openpyxl.Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)  # Remove default sheet

    for sheet_name, rows in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
